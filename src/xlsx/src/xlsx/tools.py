import json

import pandas as pd
from openpyxl import Workbook, load_workbook

from . import mcp
from .recalc import recalc


@mcp.tool()
def read_excel(file_path: str, sheet_name: str = "") -> str:
    """Read an Excel file and return its contents as markdown table.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name to read (default: first sheet)

    Returns:
        Markdown formatted table of the data
    """
    if sheet_name:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        df = pd.read_excel(file_path)
    return df.to_markdown(index=False)


@mcp.tool()
def create_excel(file_path: str, data: str, sheet_name: str = "Sheet1") -> str:
    """Create a new Excel file with data.

    Args:
        file_path: Path to output Excel file
        data: CSV formatted data to write
        sheet_name: Name of the sheet (default: 'Sheet1')

    Returns:
        Success message
    """
    from io import StringIO

    df = pd.read_csv(StringIO(data))
    df.to_excel(file_path, sheet_name=sheet_name, index=False)
    return f"Created {file_path} with {len(df)} rows"


@mcp.tool()
def write_cell(file_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Write a value or formula to a specific cell.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet
        cell: Cell address (e.g., 'A1', 'B5')
        value: Value or formula to write (formulas start with '=')

    Returns:
        Success message
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    ws[cell] = value
    wb.save(file_path)
    return f"Wrote '{value}' to {sheet_name}!{cell}"


@mcp.tool()
def recalculate(file_path: str, timeout: int = 30) -> str:
    """Recalculate all formulas in Excel file and check for errors.

    Args:
        file_path: Path to Excel file
        timeout: Maximum time to wait for recalculation (default: 30 seconds)

    Returns:
        JSON formatted result with error details
    """
    result = recalc(file_path, timeout)
    return json.dumps(result, indent=2)


@mcp.tool()
def get_sheet_names(file_path: str) -> str:
    """Get list of sheet names in an Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        Comma-separated list of sheet names
    """
    wb = load_workbook(file_path, read_only=True)
    sheets = ", ".join(wb.sheetnames)
    wb.close()
    return sheets


@mcp.tool()
def add_sheet(file_path: str, sheet_name: str) -> str:
    """Add a new sheet to an Excel file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name for the new sheet

    Returns:
        Success message
    """
    wb = load_workbook(file_path)
    wb.create_sheet(sheet_name)
    wb.save(file_path)
    return f"Added sheet '{sheet_name}' to {file_path}"


@mcp.tool()
def convert_to_csv(file_path: str, output_file: str, sheet_name: str = "") -> str:
    """Convert an Excel sheet to CSV.

    Args:
        file_path: Path to Excel file
        output_file: Path to output CSV file
        sheet_name: Sheet name to convert (default: first sheet)

    Returns:
        Success message
    """
    if sheet_name:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        df = pd.read_excel(file_path)
    df.to_csv(output_file, index=False)
    return f"Converted {file_path} to {output_file}"
